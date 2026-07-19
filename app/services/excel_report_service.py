from pathlib import Path

from openpyxl import Workbook, load_workbook

REPORT_DIR = Path("reports")


def seconds_to_hms(seconds: float) -> str:
    seconds = int(seconds)

    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60

    return f"{hours:02}:{minutes:02}:{secs:02}"


def update_shift_report(
    machine_name: str,
    shift_date,
    shift_name: str,
    up_seconds: float,
    down_seconds: float,
    efficiency: float,
):
    """
    Creates/updates:

    reports/
        MACHINE_NAME/
            shift_report.xlsx
    """

    machine_folder = REPORT_DIR / machine_name
    machine_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook_path = machine_folder / "shift_report.xlsx"

    if workbook_path.exists():
        workbook = load_workbook(workbook_path)

    else:
        workbook = Workbook()

        workbook.remove(workbook.active)

    sheet_name = shift_date.strftime("%Y-%m-%d")

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

    else:
        sheet = workbook.create_sheet(sheet_name)

        sheet.append(
            [
                "Shift",
                "Uptime",
                "Downtime",
                "Efficiency (%)",
            ]
        )

    row_number = None

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == shift_name:
            row_number = row
            break

    if row_number is None:
        row_number = sheet.max_row + 1

    sheet.cell(row=row_number, column=1).value = shift_name
    sheet.cell(row=row_number, column=2).value = seconds_to_hms(up_seconds)
    sheet.cell(row=row_number, column=3).value = seconds_to_hms(down_seconds)
    sheet.cell(row=row_number, column=4).value = round(
        efficiency,
        2,
    )

    workbook.save(workbook_path)