from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def format_duration(seconds):
    seconds = int(seconds or 0)

    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60

    return f"{hours:02}:{minutes:02}:{seconds:02}"


def create_excel_report(history):

    workbook = Workbook()

    # -----------------------------
    # Daily Performance Sheet
    # -----------------------------
    daily_sheet = workbook.active
    daily_sheet.title = "Daily Performance"

    headers = [
        "Date",
        "Uptime",
        "Downtime",
        "Efficiency (%)",
    ]

    for col, header in enumerate(headers, start=1):
        cell = daily_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    row_index = 2

    for row in history["daily"]:

        daily_sheet.cell(
            row=row_index,
            column=1,
            value=row["date"],
        )

        daily_sheet.cell(
            row=row_index,
            column=2,
            value=format_duration(
                row["uptime_seconds"]
            ),
        )

        daily_sheet.cell(
            row=row_index,
            column=3,
            value=format_duration(
                row["downtime_seconds"]
            ),
        )

        daily_sheet.cell(
            row=row_index,
            column=4,
            value=row["efficiency_pct"],
        )

        row_index += 1

    # -----------------------------
    # Shift Performance Sheet
    # -----------------------------
    shift_sheet = workbook.create_sheet(
        "Shift Performance"
    )

    headers = [
        "Date",
        "Shift",
        "Uptime",
        "Downtime",
        "Efficiency (%)",
        "Status",
    ]

    for col, header in enumerate(headers, start=1):
        cell = shift_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    row_index = 2

    for row in history["shifts"]:

        shift_sheet.cell(
            row=row_index,
            column=1,
            value=row["date"],
        )

        shift_sheet.cell(
            row=row_index,
            column=2,
            value=row["shift"],
        )

        shift_sheet.cell(
            row=row_index,
            column=3,
            value=format_duration(
                row["uptime_seconds"]
            ),
        )

        shift_sheet.cell(
            row=row_index,
            column=4,
            value=format_duration(
                row["downtime_seconds"]
            ),
        )

        shift_sheet.cell(
            row=row_index,
            column=5,
            value=row["efficiency_pct"],
        )

        shift_sheet.cell(
            row=row_index,
            column=6,
            value="Final"
            if row["is_final"]
            else "In Progress",
        )

        row_index += 1

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output